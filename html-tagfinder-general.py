from openpyxl import load_workbook
from openpyxl import Workbook
import bs4
from bs4 import BeautifulSoup
from urllib.request import urlopen
from bs4 import Comment
import datetime
import re
from tkinter import filedialog
from tkinter import *
from tkinter import messagebox

fileopen=filedialog.askopenfilename(initialdir="",title = "Select file",filetypes = (("Excel files","*.xlsx"),("CSV files","*.csv"),("all files","*.*")))
wb=load_workbook(fileopen)
sheet=wb.active
max_row=sheet.max_row
max_column=sheet.max_column
now = datetime.datetime.now()
print('enter the name of the html tag you wish to search for, with the < and > symbols excluded.')
search_tag=input()
print('Working...')
sheet['A1'].value='URL'
sheet['B1'].value='No of <'+search_tag+ '> tags'

for i in range(2,max_row+1):
    currenturl=sheet['A'+str(i)].value
    try:
        page=urlopen(currenturl)
        soup=BeautifulSoup(page, features="lxml")
        comments=soup.find_all(string=lambda text:isinstance(text,Comment))
        matchlist=soup.find_all(search_tag)
        matchcount=0
        for x in range(len(matchlist)):
            matchcount=matchcount+1
        sheet.cell(row=i, column=1).value=currenturl
        sheet.cell(row=i, column=2).value=matchcount

    except:
        sheet.cell(row=i, column=1).value=currenturl
        sheet.cell(row=i, column=2).value='ERROR'

    
filesave=filedialog.asksaveasfilename(initialdir="",title = "Save output",filetypes = (("Excel files","*.xlsx"),("all files","*.*")),defaultextension='.xlsx')
wb.save(filesave)
messagebox.showinfo('Alert','Done!')


